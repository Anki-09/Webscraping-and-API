import flask
import PyPDF2
import json
import openpyxl
from openpyxl.cell import get_column_letter
app=flask.Flask(__name__)                     #instantiate flask
@app.route('/pdf',methods=['GET'])            #create route /pdf with methods allowed - 'GET'
def home():
    text=''
    file=open('demo.pdf', 'rb')               #open demo.pdf with read binary mode
    read=PyPDF2.PdfFileReader(file)           #To read the PDF
    for i in range(read.numPages):            #Repeat for all the pages in demo.pdf
        page=read.getPage(i)                  #get each page
        text=text+page.extractText()          #Extract data from each page and append it to the string 'text'
    file.close()                              #close demo.pdf
    return json.dumps(text.replace('\n',''))  #return the data in json format
@app.route('/pdf',methods=['POST'])           #route /pdf with methods allowed - 'POST'
def watermark():
    data=flask.request.get_data()             #get the data entered by the user i.e name of the watermark pdf
    try:
      wmfile=open(data,'rb')                  #open the watermark pdf
    except:                                   #catch exception incase of error in opening the pdf
      return "Couldn't open file "+data
    wmread=PyPDF2.PdfFileReader(wmfile)       #to read the watermark pdf
    file = open('demo.pdf', 'rb')             #open demo.pdf in read binary mode
    read = PyPDF2.PdfFileReader(file)         #to read demo.pdf
    write= PyPDF2.PdfFileWriter()             #to write to the new pdf
    wmpage=wmread.getPage(0)                  #get the first page of watermark pdf (watermark page)
    for i in range(read.numPages):            #iterate through all the pages of demo.pdf
        page=read.getPage(i)                  #get each page demo.pdf
        page.mergePage(wmpage)                #merge with watermark page
        write.addPage(page)                   #To add merged page to new pdf
    outfile=open('output.pdf','wb')           #create and/or open the output.pdf in write binary mode
    write.write(outfile)                      #write all the new pdf content to output.pdf
    wmfile.close()                            #close watermark file
    file.close()                              #close demo.pdf
    outfile.close()                           #close output.pdf
    return "Merged pdf is output.pdf"
@app.route('/events', methods=['GET'])       #route /events with allowed method -GET
def events():
    dict={}
    list=[]
    wb=openpyxl.load_workbook('event.xlsx')  #load excel file -event.xlsx
    sheet=wb.get_sheet_by_name('Sheet1')     #get sheet named Sheet1
    for i in range(2,sheet.get_highest_row()+1):       #iterate through rows
        for r in range(1,sheet.get_highest_column()+1):  #iterate through columns
            key=sheet[get_column_letter(r)+'1'].value  #consider first row values as key values of dictionary
            value=str(sheet[get_column_letter(r)+str(i)].value)  #get data from remaining rows(one row per iteration) and store as values of dictionary corresponding to the keys
            dict.__setitem__(key,value)       #Add new key-value pair to the dictionary dict
        list.append(dict)                     #Add each dictionary to list(list of dictionary)
        dict={}                               #Reset dict to null
    return json.dumps(list)                   #return list of dictionary in json format
@app.route('/events/<webinar>', methods=['GET'])  #route events/webinarName with allowed method-GET
def eventsweb(webinar):
    dict={}
    wb = openpyxl.load_workbook('event.xlsx')    #load excel file -event.xlsx
    sheet = wb.get_sheet_by_name('Sheet1')       #get sheet named Sheet1
    for i in range(2,sheet.get_highest_row()+1):           #iterate through each row
        if webinar==sheet['A'+str(i)].value:     #if webinar matches any value in the rows of first column
            for r in range(1, sheet.get_highest_column() + 1):  #iterate through all the columns of matched row
                key = sheet[get_column_letter(r) + '1'].value  #consider first row values as key values of dictionary
                value = str(sheet[get_column_letter(r) + str(i)].value)  #get values from the matched row
                dict.__setitem__(key, value)     #Add key-value pair to the dictionary dict
            return json.dumps(dict)              #return dictionary dict in json format
    return "Webinar not found"                   #if mentioned webinarName doesn't match any values in excel file
@app.route('/events', methods=['POST'])         #route /events with methods allowed - 'POST'
def eventspost():
    event = json.loads(flask.request.get_data())  #store the user entered json data
    wb = openpyxl.load_workbook('event.xlsx')     #load excel file -event.xlsx
    sheet = wb.get_sheet_by_name('Sheet1')        #get sheet named Sheet1
    row=str(sheet.get_highest_row()+1)                      #get the row number to which new data is inserted
    sheet['A'+row]=event["Event"]                 #store respective data
    sheet['B' +row] = event["Presenter"]
    sheet['C' +row] = event["Venue"]
    sheet['D' +row] = event["Date-Time"]
    wb.save('event.xlsx')                        #save excel file
    return "Successfully added event"
@app.route('/events', methods=['PUT'])           #route /events with methods allowed - 'PUT'
def eventsput():
    dict = json.loads(flask.request.get_data())  #store the user entered json data
    wb = openpyxl.load_workbook('event.xlsx')    #load excel file -event.xlsx
    sheet = wb.get_sheet_by_name('Sheet1')       #get sheet named Sheet1
    for i in range(2,sheet.get_highest_row()+1):           #iterate through each row
        if dict["Event"]==sheet['A'+str(i)].value: #if the event value entered matches any row value of the first column
            sheet['B'+str(i)] = dict["Presenter"]  #update other column values of the matched row with user entered values
            sheet['C'+str(i)] = dict["Venue"]
            sheet['D'+str(i)] = dict["Date-Time"]
            wb.save('event.xlsx')                  #save excel file
            return "Updated details of event:"+dict["Event"]
    return "Event unavailable to update...Try inserting the data instead of updating" #if event value doesn't match any value in excel file
app.run()                                     #run the flask application
